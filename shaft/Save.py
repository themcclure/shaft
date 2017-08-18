"""
Taking a list of Officials and saving the summary to an Excel sheet
"""
__author__ = 'hammer'

from shaft import Official
from shaft import Game
from shaft import __version__
from Offical import sort_by_role

import re
import os
import datetime
from operator import attrgetter
from dateutil import relativedelta
from openpyxl import load_workbook
from openpyxl import utils
from openpyxl import Workbook

import config
assns = config.assns
types = config.types
roles = config.roles
ref_roles = config.ref_roles
nso_roles = config.nso_roles
nso_family = config.nso_family


def create_results(file_name, officials, model):
    """
    create the Excel file summarizing the officials, given the chosen weighting model
    :param file_name: the name of the file to output
    :param officials: the list of officials
    :param model: the weighting model to use
    :return: the excel object (for now)
    """
    # TODO: add in autofilters
    order = 0
    wb = Workbook()
    page1 = wb.active
    page1.title = "Applicants"

    header = officials[0].get_summary_header()
    wb['Applicants'].append(header)
    #wb['applicants'].auto_filter.ref = 'A1:BH1'

    # print summary of entire list, sorted by name
    for off in sorted(officials, key=attrgetter('name')):
        print off
        off.get_summary(model.name)
        wb['Applicants'].append(off.get_summary(model.name))

    # go through each ref role:
    for r in ref_roles:
        order += 1
        wb.create_sheet(order, r)
        wb[r].append(officials[0].get_role_header())
        for o in sort_by_role(officials, r, model.name):
            wb[r].append(o)

    # go through each nso role:
    for r in nso_roles:
        order += 1
        wb.create_sheet(order, r)
        wb[r].append(officials[0].get_role_header())
        for o in sort_by_role(officials, r, model.name):
            wb[r].append(o)

    # write out the weighting table used as the last tab, for comparison purposes and some metadata
    order += 1
    wb.create_sheet(order, 'model')
    wb['model'].append(['Weighting Model: ','','(generated by SHAFT ' + str(__version__) + ')'])
    wb['model'].append(['Model Name: ', model.name])
    wb['model'].append(['CH Uplift:', 'This is applied to boost the value of CH over H positions, as being more desirable for tournaments'])
    wb['model'].append([model.ch_uplift])
    wb['model'].append(['Secondary Role Factor:', 'This applied to positions in the "Secondary Position" column'])
    wb['model'].append([model.secondary_weight])
    #wb['model'].append(['Tertiary Role Factor:', model.tertiary_weight])
    #wb['model'].append(['This is for games in the "Other" tab (currently unused)'])
    wb['model'].append(['Age Decay Factor:', 'Each number represents the weighting applied to a game, depending how many years old the game is'])
    wb['model'].append(['Age:'] + range(len(model.decay)) + ['+'])
    wb['model'].append(['Factor:'] + model.decay + [model.decay[-1]])

    # print out each association in the model
    for assn in sorted(model.wgt, reverse=True):
        wb['model'].append(['Association: ', assn])

        # print out each game type in the assocation model
        # TODO: OPTIONAL: figure out how to from highest weight to lowest, because this doesn't work:
        # for t in sorted(model.wgt[assn], key=lambda val: val[1], reverse=True):
        for t in model.wgt[assn]:
            wb['model'].append(['Type: ', t, model.wgt[assn][t]])

    # save the file
    wb.save(file_name)


def create_rejects(file_name, rejects):
    '''
    Creates a file with the rejected file names, and the reson for rejection
    :param file_name: filename to write to
    :param rejects: list of tuples (history file, reject reason)
    :return: None
    '''

    # TODO: set up the file so it reads in existing file, and write out the new file, including manually caught errors (like permissions or something)

    # set up the rejects file
    wb = Workbook()
    page1 = wb.active
    page1.title = "Unprocessed Applicants"
    wb['Unprocessed Applicants'].append(['History Document', 'Error'])

    for r in rejects:
        wb['Unprocessed Applicants'].append(r)

    # save the file
    wb.save(file_name)


def create_raw_results(file_name, officials, model):
    """
    create the Excel file giving a raw dump of the officials, given the chosen weighting model
    :param file_name: the name of the file to output
    :param officials: the list of officials
    :param model: the weighting model to use
    :return: the excel object (for now)
    """
    order = 0
    wb = Workbook()
    page1 = wb.active
    page1.title = "Applicants (RAW)"

    # header row for all the game data, smooshed with the core official data
    header = []
    header.append('name')
    header.append('refcert')
    header.append('ref_tally')
    header.append('nsocert')
    header.append('nso_tally')
    wb['Applicants (RAW)'].append(header)

    wb.create_sheet(1,'Games (RAW)')
    header = []
    header.append('name')
    header.append('assn')
    header.append('age')
    header.append('date')
    header.append('type')
    header.append('role')
    header.append('value')
    header.append('event')
    header.append('primacy')
    wb['Games (RAW)'].append(header)

    # print summary of entire list, sorted by name
    for off in sorted(officials, key=attrgetter('name')):
        # print the official info in the applicants tab
        print off
        official_denormalized = []
        official_denormalized.append(off.name)
        official_denormalized.append(off.refcert)
        official_denormalized.append(off.ref_tally)
        official_denormalized.append(off.nsocert)
        official_denormalized.append(off.nso_tally)
        wb['Applicants (RAW)'].append(official_denormalized)

        # print the game data in the games tab, official's name is the joining key
        for j in off.games:
            game_denormalized = []
            game_denormalized.append(off.name)
            game_denormalized.append(j.assn)
            game_denormalized.append(j.age)
            game_denormalized.append(j.date)
            game_denormalized.append(j.type)
            game_denormalized.append(j.role)
            game_denormalized.append(model.weight(j))
            #game_denormalized.append(j.event)
            game_denormalized.append('')
            game_denormalized.append(j.primacy)
            #print "name = %s, role = %s, on %s" % (game_denormalized[0], game_denormalized[5], game_denormalized[3])
            wb['Games (RAW)'].append(game_denormalized)

    # save the file
    wb.save(file_name)
